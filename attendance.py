import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
import os
import serial
import time
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import serial.tools.list_ports  # For detecting available serial ports

# Function to automatically find the serial port for the ESP32/Arduino
def find_serial_port():
    ports = list(serial.tools.list_ports.comports())  # List all serial ports
    for port, desc, hwid in ports:z
        if 'USB' in desc or 'ACM' in desc:  # Matching ports based on description
            print(f"Found device on port {port}")
            return port
    print("No suitable serial port found.")
    return None

# Set up serial communication automatically using the found port
serial_port = find_serial_port()
if serial_port:
    ser = serial.Serial(serial_port, 9600, timeout=1)
    time.sleep(2)  # Allow time for the device to initialize
else:
    print("Error: Unable to find a serial port.")
    exit(1)

# Load the master workbook to map fingerprint IDs to names
try:
    master_wb = openpyxl.load_workbook("C:/Users/pravin/OneDrive/Desktop/Documents/alcohol.xlsx")
    master_sheet = master_wb.active
except FileNotFoundError:
    print("Error: Master Excel file not found.")
    exit(1)

# Load the attendance workbook and select the active sheet
try:
    wb = openpyxl.load_workbook("C:/Users/pravin/OneDrive/Desktop/Documents/attendance.xlsx")
    sheet = wb.active
except FileNotFoundError:
    print("Error: Attendance Excel file not found.")
    exit(1)

# Assign permanent column headers if they don't already exist
def set_column_headings():
    headers = ["Date", "Time", "ID", "Name", "ALC Level", "Attendance Status", "Test Status"]
    for col_num, header in enumerate(headers, start=1):
        if sheet.cell(row=1, column=col_num).value != header:
            sheet.cell(row=1, column=col_num, value=header)
    wb.save("C:/Users/pravin/OneDrive/Desktop/Documents/attendance.xlsx")

# Function to find the name for a given fingerprint ID from the master Excel file
def get_name_from_master(finger_id):
    for row in master_sheet.iter_rows(min_row=2, values_only=True):  # Assuming IDs start from the second row
        if str(row[0]) == finger_id:  # ID expected in column A
            return row[1]  # Name expected in column B
    return "Unknown User"

# Define fill colors for attendance status and test status
present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green for "Present"
absent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Light red for "Absent"
ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")       # Light green for "OK"
ng_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")       # Light red for "NG"

# Function to send an email if the alcohol level is above the threshold
def send_email(subject, body, recipient_emails, attachment_path=None):
    sender_email = "praveentansam@gmail.com"  # Replace with your email
    sender_password = "your-app-password"  # Replace with your app password
    display_name = "Transport_Team Rnaipl"

    msg = MIMEMultipart()
    msg["From"] = f"{display_name} <{sender_email}>"
    msg["To"] = ", ".join(recipient_emails)
    msg["Subject"] = subject

    msg.attach(MIMEText(body, "plain"))

    # Attach the file if path is provided
    if attachment_path and os.path.exists(attachment_path):
        with open(attachment_path, "rb") as f:
            file_attachment = MIMEImage(f.read())
            file_attachment.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(attachment_path)}"')
            msg.attach(file_attachment)

    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, recipient_emails, msg.as_string())
        server.quit()
        print("Email sent successfully to supervisor!")
    except Exception as e:
        print(f"Failed to send email: {e}")

# Function to log attendance in the attendance Excel file and send email if alcohol level exceeds threshold
def log_attendance(finger_id, name, alcohol_level):
    row = sheet.max_row + 1
    current_time = datetime.now()
    attendance_status = 'Present' if alcohol_level < 3500 else 'Absent'
    test_status = 'OK' if alcohol_level < 3500 else 'NG'

    # Fill in attendance details
    sheet.cell(row=row, column=1, value=current_time.strftime("%Y-%m-%d"))
    sheet.cell(row=row, column=2, value=current_time.strftime("%H:%M:%S"))
    sheet.cell(row=row, column=3, value=finger_id)
    sheet.cell(row=row, column=4, value=name)
    sheet.cell(row=row, column=5, value=alcohol_level)

    attendance_cell = sheet.cell(row=row, column=6, value=attendance_status)
    attendance_cell.fill = present_fill if attendance_status == 'Present' else absent_fill

    test_status_cell = sheet.cell(row=row, column=7, value=test_status)
    test_status_cell.fill = ok_fill if test_status == 'OK' else ng_fill

    try:
        wb.save("C:/Users/pravin/OneDrive/Desktop/Documents/attendance.xlsx")
        print(f"Attendance logged for Fingerprint ID {finger_id} ({name}) with Alcohol Level {alcohol_level}")

        # Send email if alcohol level exceeds 3500
        if alcohol_level >= 3500:
            subject = f"Alcohol test NG Alert for {name}"
            body = f"Alert: {name}\n (ID: {finger_id})\n has an alcohol level of {alcohol_level}\n at {current_time.strftime('%Y-%m-%d %H:%M:%S')}\n \n\n Take Immediate attention required."
            recipient_emails = ["seshanjai@gmail.com"]  # Replace with supervisor's email
            send_email(subject, body, recipient_emails)

    except PermissionError:
        print("Error: Permission denied when saving the Excel file. Make sure it's not open elsewhere.")

# Set up column headings at the start of the script
set_column_headings()

finger_id, name, alcohol_level = None, None, None  # Variables to store data between messages

while True:
    if ser.in_waiting > 0:
        data = ser.readline().decode('utf-8').strip()
        print("Received from Arduino:", data)

        if "Fingerprint ID found:" in data:
            finger_id = data.split(":")[1].strip()
            print(f"Fingerprint ID: {finger_id}")
            name = get_name_from_master(finger_id)
            print(f"User Name: {name}")

        elif "ID:" in data and "ALC:" in data:
            parts = data.split(",")
            finger_id = parts[0].split(":")[1].strip()
            alcohol_level = int(parts[1].split(":")[1].strip())
            print(f"Fingerprint ID: {finger_id}, Alcohol Level: {alcohol_level}")

        elif "Status: Present, OK" in data:
            if finger_id and name and alcohol_level is not None:
                log_attendance(finger_id, name, alcohol_level)
            else:
                print("Error: Missing data for logging attendance.")

        elif "Status: Absent" in data:
            if finger_id and name and alcohol_level is not None:
                log_attendance(finger_id, name, alcohol_level)
            else:
                print("Error: Missing data for logging attendance.")

    time.sleep(1)  # Delay to avoid flooding
